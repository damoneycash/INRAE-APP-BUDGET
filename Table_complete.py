import streamlit as st
import streamlit as st
from numpy import *
from pandas import *
from openpyxl import *
import streamlit as st
import datetime
from os.path import splitext, basename

def table_complete1(drap1, drap):
    e = str(datetime.datetime.today())
    e1 = e.split(" ")[0]
    e2 = e1.split("-")[0]
    year = int(e2)+1

    L = []
    for i in drap1["F"]:
        L.append(i.value)
    L.pop(0)
    L.pop(0)
    L.pop(0)
#print(L)
    P = []
    for i in range(len(L)):
        if str(L[i]) == 'None':
            L[i] = 1
        elif str(L[i]) == 'Terminé':
            L[i] = datetime.datetime(2020, 11, 11, 0, 0)
        elif str(L[i]) == 'PROLONGE':
            L[i] = 1
        elif str(L[i]) == ' CHAQUE ANNEE':
            L[i] = datetime.datetime(year, 11, 11, 0, 0)
    W = []
    z = 0
    for i in range(len(L)):
        if L[i] == 1:
            z += 1
        else :
            W.append(L[i])
    Ldate = array(W)
#print(Ld)


#file = glob2.glob('EXCEL\*.xlsx')
#print(drap)
#print(wb[1])
#print(classeur.get_sheet_names())
#print(drap["H"])

#Get liste Scientifiques'''
    L = []
    for i in drap["H"]:
        L.append(i.value)
    L.pop(0)
    L.sort()
    Sc = array(L).unique()
#print(Sc)

#Get liste lignes analytiques
    L = []
    for i in drap["L"]:
        L.append(i.value)
    L.pop(0)
    L.sort()
    La = array(L).unique()

#Get toute la data
    a = ''
    for line in drap:
        for i in line:
        #print(str(i.value))
            a =  a + "{" + str(i.value)
#print(a)
#Split la data
    Sc_data = []
    Sc_data_final = []
    Sc_data.append(a.split("BPP"))
    V = Sc_data[0][0]
    Sc_data[0].pop(0)
#print(Sc_data[0][1])
    for j in range(len(Sc_data[0])-1):
        A = []
        a = ''
        a = str(Sc_data[0][j])
        split1 = a.split('{')
    #print(split1)
        for i in split1:
            A.append(i)
        A.pop(0)
        A.pop(len(split1)-2)
    #print(A)
        Sc_data_final.append(A)
    a = ''
    a = str(Sc_data[0][len(Sc_data)-2])
    split1 = a.split('{')
#print(str(Sc_data[0][len(Sc_data)-2]))
#print(len(Sc_data[0]))
    A = []
    for i in split1:
        A.append(i)
    A.pop(0)
#print(A)
    Sc_data_final.append(A)
#for k in range(len(Sc_data_final)):
    #print(Sc_data_final[k])

#Répartir la data pour chaque scientifique
    K = []
    L = []
    Z = []
    for i in range(len(Sc)):
        K.append(L)
#print(K)
    J = []
    w = 0
    for i in range(len(Sc)): #len(Sc)
        L = []
        e = 0
        for k in range(len(Sc_data_final)):
            if str(Sc_data_final[k][6]) == str(Sc[i]):
            #K[i].append(Sc_data_final[k])
                L.append(e)
            #print(Sc_data_final[e][6])   #pass      
            e += 1
        for j in L:
        #print(j)
            J.append(str(Sc_data_final[int(j)]))
        Z.append(J)
        L = []
        J = []
#print(Z[7:20])
    L1, L2, L3 = [], [], []
    for i in range(len(Z)): 
        L2 = []
        for j in Z[i]:
            L1 = []
        #print(j)
            r = ''
            split2 = j.split("', '")
            for z in range(2, len(split2[0])):
                r += split2[0][z]
            split2[0] = r    
            r = ''
            for z in range(len(split2[len(split2)-1])-2):
                r += split2[len(split2)-1][z]
            split2[len(split2)-1] = r
        #print(split2)
            for k in split2:
            #print(k)
            #print(k)
                L1.append(k)
            L2.append(L1)
        L3.append(L2)
#print(L3[0][0][0])
    Sc_data = L3
    L = []
    for i in range(len(Sc_data)):
        for j in range(len(Sc_data[i])):
            L.append(Sc_data[i][j][7])
    Li = array(L).unique()
#print(Li)


#print(Sc_data)
    key = []
    for i in range(len(Li)):
        B = []
        B.append(Ldate[i])
        B.append(Li[i])
        key.append(B)
#print(key)

    
    G = []
    for i in range(len(Sc_data)):
        for j in range(len(Sc_data[i])):
            for k in range(len(key)):
                if key[k][1] == Sc_data[i][j][7]:
                    Sc_data[i][j].append(key[k][0])
                    G.append(key[k][0])
    for i in range(len(Sc_data)):
        for j in range(len(Sc_data[i])):
        #Sc_data[i][j] = Sc_data[i][j][:16] + Sc_data[i][j][17:]
            if len(Sc_data[i][j]) != len(Sc_data[0][0]):
                Sc_data[i][j].append(datetime.datetime(999, 1, 1, 0, 0))
                G.append(datetime.datetime(999, 1, 1, 0, 0))

#print(Sc_data)
    L2 = []
    for i in range(len(Sc_data)):
        L = []
        for j in range(len(Sc_data[i])):
            L.append(Sc_data[i][j][16])
        M = []
        for r in L:
            a = str(r)
            split1 = a.split(" ")[0]
            M.append(split1)
        #print(M)
        D = M.copy()
        M.sort()
        L1 = []
        #print(D)
        #print(M)
        for z in range(len(M)):
            for l in range(len(D)):
                if M[z] == D[l]:
                    Sc_data[i][l][16] = M[z]
                    L1.append(Sc_data[i][l])
                
    #print(L1)        
        L2.append(L1)
    Sc_data = L2
    unique_data = []
    seen = set()
    for scientist_data in Sc_data:
        unique_scientist_data = []
        for row in scientist_data:
            if tuple(row) not in seen:
                seen.add(tuple(row))
                unique_scientist_data.append(row)
        unique_data.append(unique_scientist_data)

    Sc_data = unique_data
    Table_final = []
    for i in range(len(Sc_data)):
        for j in range(len(Sc_data[i])):
            Table_final.append(Sc_data[i][j])
    df = DataFrame(Table_final)
    Headers = []
    V = str(V) + "Date"
    b = str(V)
    split = b.split("{")
    for i in split:
        Headers.append(i)
    Headers.pop(0)
    Headers.pop(0)
    df.columns=Headers
    #st.markdown(' ')
    st.dataframe(data=df, use_container_width= True, height=700)
    st.column_config.TextColumn(width="small")
    return 

def saut_ligne(i):
    for i in range(int(i)):
        st.write("")
    return
def main():
    st.write('<img class="Logo-etat" src="https://www.inrae.fr/themes/custom/inrae_socle/public/images/etat_logo.svg" alt="République française" width="138" height="146">',
             '<img class="Logo-site" src="https://www.inrae.fr/themes/custom/inrae_socle/logo.svg" alt="INRAE">',
             "<h1 style='text-align: center; color : aqua'>Traitement de donnée Excel : </h1>",
             unsafe_allow_html=True)
    saut_ligne(2)
    '''
    file = glob2.glob('EXCEL\*.xlsx')
    names = []
    for i in file:
        names.append(splitext(basename(i))[0])
    Sc_dico = {}
    z = 0
    for i in names:
        Sc_dico[i] = z
        z += 1
    option = st.selectbox("Choisissez le fichier excel ⬇️", names)
    m = Sc_dico[option]
    saut_ligne(2)
    st.title(":blue[Table complète étudiée 📊] ")
    saut_ligne(3)
    classeur = load_workbook(file[m]) #file
    wb = classeur.sheetnames
    wa = wb[1]
    drap = classeur[wa]
    '''
    uploaded_file = st.file_uploader("Téléversez un fichier Excel", type=["xlsx", "xls"])
    if uploaded_file is not None:
        try:
            classeur = load_workbook(uploaded_file)
            wb = classeur.sheetnames
            wa = wb[0]
            wc = wb[1]
            drap1 = classeur[wa]
            drap = classeur[wc]
            table_complete1(drap1, drap)
            # Lecture du fichier Excel avec Pandas
            
           
        except UnboundLocalError:
            # Afficher "Hello World" en cas d'erreur
            st.write("Hello World")
    return
